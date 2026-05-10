#!/usr/bin/env python3
"""Fitness tracker CLI: init schema + sync from Excel + goal-check."""
import argparse
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
DB_PATH = ROOT / "fitness.db"
EXCEL_PATH = Path(r"H:\My Drive\8 - Fitness\Growth V3.xlsx")

TARGET_WEIGHT_KG = 73.0
TARGET_CHEST_WEIGHT_KG = 100.0
TARGET_CHEST_REPS = 15
TARGET_RUN_DISTANCE_KM = 10.0
TARGET_PACE_SEC_PER_KM = 300  # 5:00/km
STALENESS_DAYS = 7

GREEN = "\033[92m"
RED = "\033[91m"
YELLOW = "\033[93m"
DIM = "\033[2m"
RESET = "\033[0m"


def connect():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def create_schema(conn):
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS chest_press (
            date       TEXT PRIMARY KEY,
            weight_kg  REAL NOT NULL,
            reps       INTEGER NOT NULL
        );
        CREATE TABLE IF NOT EXISTS body_weight (
            date       TEXT PRIMARY KEY,
            weight_kg  REAL NOT NULL
        );
        CREATE TABLE IF NOT EXISTS running (
            date              TEXT PRIMARY KEY,
            distance_km       REAL NOT NULL,
            duration_seconds  INTEGER NOT NULL,
            elevation_gain_m  REAL NOT NULL
        );
        """
    )
    conn.commit()


def to_iso_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def parse_run_duration(t):
    # Excel cell is 'h:mm' formatted but the value is mm:ss elapsed
    # (e.g. time(19, 36) = 19 min 36 sec). Whole-second resolution;
    # runs >= 60min require reformatting the cell to [mm]:ss.
    if t is None:
        return None
    return t.hour * 60 + t.minute


def sync_from_excel(conn):
    if not EXCEL_PATH.exists():
        print(f"WARN: Excel not found at {EXCEL_PATH}; skipping sync", file=sys.stderr)
        return 0
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    n = 0

    if "Chest" in wb.sheetnames:
        for row in wb["Chest"].iter_rows(min_row=3, values_only=True):
            d = to_iso_date(row[1]) if len(row) > 1 else None
            wt = row[2] if len(row) > 2 else None
            reps = row[3] if len(row) > 3 else None
            if d is None or wt is None or reps is None:
                continue
            conn.execute(
                "INSERT INTO chest_press(date, weight_kg, reps) VALUES (?,?,?) "
                "ON CONFLICT(date) DO UPDATE SET weight_kg=excluded.weight_kg, reps=excluded.reps",
                (d, float(wt), int(reps)),
            )
            n += 1

    if "Weight" in wb.sheetnames:
        for row in wb["Weight"].iter_rows(min_row=3, values_only=True):
            d = to_iso_date(row[1]) if len(row) > 1 else None
            wt = row[2] if len(row) > 2 else None
            if d is None or wt is None:
                continue
            conn.execute(
                "INSERT INTO body_weight(date, weight_kg) VALUES (?,?) "
                "ON CONFLICT(date) DO UPDATE SET weight_kg=excluded.weight_kg",
                (d, float(wt)),
            )
            n += 1

    if "Run" in wb.sheetnames:
        for row in wb["Run"].iter_rows(min_row=3, values_only=True):
            d = to_iso_date(row[1]) if len(row) > 1 else None
            dist = row[2] if len(row) > 2 else None
            dur = parse_run_duration(row[3]) if len(row) > 3 else None
            elev = row[4] if len(row) > 4 else 0
            if d is None or dist is None or dur is None:
                continue
            conn.execute(
                "INSERT INTO running(date, distance_km, duration_seconds, elevation_gain_m) "
                "VALUES (?,?,?,?) "
                "ON CONFLICT(date) DO UPDATE SET "
                "  distance_km=excluded.distance_km, "
                "  duration_seconds=excluded.duration_seconds, "
                "  elevation_gain_m=excluded.elevation_gain_m",
                (d, float(dist), int(dur), float(elev or 0)),
            )
            n += 1

    conn.commit()
    return n


def cmd_init(args):
    if args.force and DB_PATH.exists():
        DB_PATH.unlink()
        print(f"Removed existing {DB_PATH}")
    conn = connect()
    create_schema(conn)
    print(f"Schema ready at {DB_PATH}")
    n = sync_from_excel(conn)
    print(f"Imported {n} rows from {EXCEL_PATH}")
    conn.close()


def cmd_sync(args):
    if not DB_PATH.exists():
        print(f"DB not found at {DB_PATH}. Run `python fitness.py init` first.", file=sys.stderr)
        sys.exit(1)
    conn = connect()
    n = sync_from_excel(conn)
    print(f"Synced {n} rows from {EXCEL_PATH}")
    conn.close()


def status_label(ok):
    return f"{GREEN}PASS{RESET}" if ok else f"{RED}FAIL{RESET}"


def days_old(d_iso):
    return (date.today() - date.fromisoformat(d_iso)).days


def fmt_pace(sec_per_km):
    m = int(sec_per_km // 60)
    s = int(round(sec_per_km - m * 60))
    if s == 60:
        m, s = m + 1, 0
    return f"{m}:{s:02d}/km"


def cmd_goal_check(args):
    if not DB_PATH.exists():
        print(f"DB not found at {DB_PATH}. Run `python fitness.py init` first.", file=sys.stderr)
        sys.exit(1)
    conn = connect()

    print()
    print(f"Goal status as of {date.today().isoformat()} (rolling 7-day window)")
    print()

    all_pass = True
    any_stale = False

    # 1. Body weight < 73kg
    row = conn.execute(
        "SELECT date, weight_kg FROM body_weight ORDER BY date DESC LIMIT 1"
    ).fetchone()
    if row:
        d, w = row
        ok = w < TARGET_WEIGHT_KG
        age = days_old(d)
        stale = age > STALENESS_DAYS
        all_pass = all_pass and ok and not stale
        any_stale = any_stale or stale
        flag = f" {YELLOW}[stale {age}d]{RESET}" if stale else ""
        print(
            f"  {status_label(ok)}  Body weight     {w:5.1f} kg   "
            f"{DIM}on {d}, target <{TARGET_WEIGHT_KG:.0f} kg{RESET}{flag}"
        )
    else:
        all_pass = False
        print(f"  {RED}NODATA{RESET}  Body weight")

    # 2. Chest press 100kg x 15+ reps
    row = conn.execute(
        "SELECT date, weight_kg, reps FROM chest_press ORDER BY date DESC LIMIT 1"
    ).fetchone()
    if row:
        d, w, r = row
        ok = w >= TARGET_CHEST_WEIGHT_KG and r >= TARGET_CHEST_REPS
        age = days_old(d)
        stale = age > STALENESS_DAYS
        all_pass = all_pass and ok and not stale
        any_stale = any_stale or stale
        flag = f" {YELLOW}[stale {age}d]{RESET}" if stale else ""
        print(
            f"  {status_label(ok)}  Chest press     {w:5.1f} kg x {r:>2} reps   "
            f"{DIM}on {d}, target {TARGET_CHEST_WEIGHT_KG:.0f}kg x{TARGET_CHEST_REPS}+{RESET}{flag}"
        )
    else:
        all_pass = False
        print(f"  {RED}NODATA{RESET}  Chest press")

    # 3. 10km @ <=5:00/km
    row = conn.execute(
        "SELECT date, distance_km, duration_seconds FROM running ORDER BY date DESC LIMIT 1"
    ).fetchone()
    if row:
        d, dist, dur = row
        pace = dur / dist
        ok = dist >= TARGET_RUN_DISTANCE_KM and pace <= TARGET_PACE_SEC_PER_KM
        age = days_old(d)
        stale = age > STALENESS_DAYS
        all_pass = all_pass and ok and not stale
        any_stale = any_stale or stale
        flag = f" {YELLOW}[stale {age}d]{RESET}" if stale else ""
        print(
            f"  {status_label(ok)}  Run             {dist:5.2f} km @ {fmt_pace(pace)}   "
            f"{DIM}on {d}, target 10km @ <={fmt_pace(TARGET_PACE_SEC_PER_KM)}{RESET}{flag}"
        )
    else:
        all_pass = False
        print(f"  {RED}NODATA{RESET}  Run")

    print()
    overall = f"{GREEN}ALL GOALS MET{RESET}" if all_pass else f"{RED}NOT YET{RESET}"
    print(f"  Overall: {overall}")
    if any_stale:
        print(f"  {YELLOW}One or more entries are older than {STALENESS_DAYS} days.{RESET}")
    print()
    conn.close()


def main():
    p = argparse.ArgumentParser(prog="fitness", description="Fitness tracker (Excel-backed SQLite store)")
    sub = p.add_subparsers(dest="cmd", required=True)

    p_init = sub.add_parser("init", help="Create schema and seed from Excel")
    p_init.add_argument("--force", action="store_true", help="Drop the existing DB first")
    p_init.set_defaults(func=cmd_init)

    p_sync = sub.add_parser("sync", help="Re-import latest values from Excel (upsert by date)")
    p_sync.set_defaults(func=cmd_sync)

    p_check = sub.add_parser("goal-check", help="Latest-entry status per goal with green/red and stale flag")
    p_check.set_defaults(func=cmd_goal_check)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
